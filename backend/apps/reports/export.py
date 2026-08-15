"""Report export — CSV, PDF, XLSX, ZIP."""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import datetime
from decimal import Decimal
from typing import Any

from django.core.files.base import ContentFile


def _safe_val(v: Any) -> Any:
    if v is None:
        return ''
    if isinstance(v, (int, float, Decimal)):
        return v
    return str(v)


def _safe_str_for_pdf(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, Decimal):
        return f'{v:,.2f}'
    return str(v)


def export_csv(headers: list[str], rows: list[list[Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_safe_val(c) for c in row])
    return buf.getvalue().encode('utf-8-sig')


def export_xlsx(headers: list[str], rows: list[list[Any]], sheet_name: str = 'Report') -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_idx, row in enumerate(rows, 2):
        ws.append([_safe_val(c) for c in row])
        for c_idx, cell in enumerate(ws[r_idx], 1):
            cell.border = border
            if isinstance(cell.value, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal='right')
                
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(title: str, headers: list[str], rows: list[list[Any]], meta: dict | None = None) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'EnterpriseTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=18,
        spaceAfter=12,
        textColor=colors.HexColor('#0f172a')
    )
    subtitle_style = ParagraphStyle(
        'EnterpriseSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#64748b')
    )
    
    story = [
        Paragraph("<b>FUTmRide</b> | Enterprise Reporting", subtitle_style),
        Spacer(1, 10),
        Paragraph(title, title_style),
        Spacer(1, 8),
    ]
    if meta:
        meta_data = [[Paragraph(f'<b>{k}:</b>', subtitle_style), Paragraph(_safe_str_for_pdf(v), subtitle_style)] for k, v in meta.items()]
        if meta_data:
            meta_table = Table(meta_data, colWidths=[120, 300])
            meta_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 16))

    data = [headers] + [[_safe_str_for_pdf(c) for c in row] for row in rows[:500]]
    if len(rows) > 500:
        data.append(['…'] * len(headers))
        data.append([f'(Truncated — {len(rows)} total rows)'] + [''] * (len(headers) - 1))

    table = Table(data, repeatRows=1)
    
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ])
    
    for col_idx, h in enumerate(headers):
        if any(keyword in h.lower() for keyword in ('amount', 'fee', 'fare', 'price', 'total', 'commission', 'payout', 'ngn', 'value')):
            table_style.add('ALIGN', (col_idx, 0), (col_idx, -1), 'RIGHT')

    table.setStyle(table_style)
    story.append(table)
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f'Generated {datetime.now().strftime("%Y-%m-%d %H:%M")} · FUTmRide Enterprise System · Strictly Confidential',
        ParagraphStyle('Footer', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=8, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER)
    ))
    doc.build(story)
    return buf.getvalue()


def export_zip(files: list[tuple[str, bytes]]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, content in files:
            zf.writestr(name, content)
    return buf.getvalue()


def build_file_content(
    fmt: str,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    meta: dict | None = None,
) -> tuple[bytes, str]:
    meta = meta or {}
    if fmt == 'csv':
        return export_csv(headers, rows), 'text/csv'
    if fmt == 'xlsx':
        return export_xlsx(headers, rows, title[:31]), (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    if fmt == 'pdf':
        return export_pdf(title, headers, rows, meta), 'application/pdf'
    if fmt == 'zip':
        inner = export_csv(headers, rows)
        return export_zip([(f'{title.replace(" ", "_")}.csv', inner)]), 'application/zip'
    return export_csv(headers, rows), 'text/csv'


def save_report_file(run, content: bytes, ext: str):
    filename = f'{run.report_key}_{run.id.hex[:8]}.{ext}'
    run.file.save(filename, ContentFile(content), save=False)
    run.file_size = len(content)
